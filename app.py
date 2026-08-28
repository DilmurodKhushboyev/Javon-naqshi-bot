import os
import re
import io
import sqlite3
from datetime import datetime

import telebot
from telebot import types
from flask import Flask, request
from openpyxl import Workbook, load_workbook

# ============ SOZLAMALAR ============
BOT_TOKEN = os.environ.get("BOT_TOKEN")
ADMIN_IDS = [int(x) for x in os.environ.get("ADMIN_IDS", "").split(",") if x.strip()]
WEBHOOK_URL = os.environ.get("WEBHOOK_URL")  # masalan: https://sizning-app.up.railway.app

# Railway Volume shu papkaga ulanadi (Volume "Mount Path" ni /data qilib sozlang)
DB_PATH = os.environ.get("DB_PATH", "/data/kutubxona.db")

bot = telebot.TeleBot(BOT_TOKEN, threaded=True)
app = Flask(__name__)


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS books (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            author TEXT,
            shelf INTEGER NOT NULL,
            row_num INTEGER NOT NULL,
            qty INTEGER,
            added_at TEXT
        )
    """)
    conn.commit()
    conn.close()


init_db()

# Admin qadam-baqadam qo'shish holatini xotirada saqlash: {user_id: {"step": ..., "data": {...}}}
user_states = {}


# ============ YORDAMCHI FUNKSIYALAR ============

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def format_book(book) -> str:
    title = book["title"] if book["title"] else "—"
    text = f"📖 <b>{title}</b>\n"
    if book["author"]:
        text += f"✍️ Muallif: {book['author']}\n"
    text += f"📍 Javon: <b>{book['shelf']}</b>  |  Qator: <b>{book['row_num']}</b>"
    if book["qty"]:
        text += f"\n🔢 Soni: {book['qty']}"
    return text


def search_books(query: str, limit: int = 10):
    """Nomi yoki muallif bo'yicha qisman, katta-kichik harfga sezgirmas qidiruv."""
    query = query.strip()
    if not query:
        return []
    like_pattern = f"%{query}%"
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM books WHERE title LIKE ? COLLATE NOCASE OR author LIKE ? COLLATE NOCASE LIMIT ?",
        (like_pattern, like_pattern, limit),
    ).fetchall()
    conn.close()
    return rows


def main_menu_keyboard(user_id: int):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(types.KeyboardButton("🔍 Kitob qidirish"))
    if is_admin(user_id):
        kb.add(types.KeyboardButton("➕ Kitob qo'shish"), types.KeyboardButton("📂 Excel yuklash"))
        kb.add(types.KeyboardButton("📋 Javon bo'yicha ko'rish"), types.KeyboardButton("📊 Statistika"))
    return kb


def cancel_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(types.KeyboardButton("❌ Bekor qilish"))
    return kb


# ============ /start ============

@bot.message_handler(commands=["start"])
def cmd_start(message):
    user_states.pop(message.from_user.id, None)
    text = (
        "📚 <b>Kutubxona botiga xush kelibsiz!</b>\n\n"
        "Kitob nomini shunchaki yozing — men uning qaysi javon va qatorda "
        "turganini topib beraman.\n\n"
        "Masalan: <i>Oʻtkan kunlar</i>"
    )
    bot.send_message(message.chat.id, text, reply_markup=main_menu_keyboard(message.from_user.id))


@bot.message_handler(commands=["cancel"])
def cmd_cancel(message):
    user_states.pop(message.from_user.id, None)
    bot.send_message(message.chat.id, "Bekor qilindi.", reply_markup=main_menu_keyboard(message.from_user.id))


# ============ ADMIN: Kitob qo'shish (qadam-baqadam) ============

@bot.message_handler(func=lambda m: m.text == "➕ Kitob qo'shish")
def add_book_start(message):
    if not is_admin(message.from_user.id):
        return
    user_states[message.from_user.id] = {"step": "title", "data": {}}
    bot.send_message(message.chat.id, "Kitob nomini yozing:", reply_markup=cancel_keyboard())


@bot.message_handler(commands=["add"])
def cmd_add(message):
    add_book_start(message)


@bot.message_handler(func=lambda m: m.text == "❌ Bekor qilish")
def handle_cancel_button(message):
    user_states.pop(message.from_user.id, None)
    bot.send_message(message.chat.id, "Bekor qilindi.", reply_markup=main_menu_keyboard(message.from_user.id))


@bot.message_handler(func=lambda m: user_states.get(m.from_user.id, {}).get("step") is not None)
def handle_add_flow(message):
    uid = message.from_user.id
    state = user_states.get(uid)
    if not state:
        return
    step = state["step"]
    data = state["data"]

    if step == "title":
        data["title"] = message.text.strip()
        state["step"] = "author"
        bot.send_message(message.chat.id, "Muallifini yozing (bilmasangiz \"-\" yozing):")

    elif step == "author":
        author = message.text.strip()
        data["author"] = "" if author == "-" else author
        state["step"] = "shelf"
        bot.send_message(message.chat.id, "Javon raqamini yozing (masalan: 3):")

    elif step == "shelf":
        if not message.text.strip().isdigit():
            bot.send_message(message.chat.id, "Iltimos, faqat raqam kiriting. Javon raqami:")
            return
        data["shelf"] = int(message.text.strip())
        state["step"] = "row"
        bot.send_message(message.chat.id, "Qator raqamini yozing (masalan: 2):")

    elif step == "row":
        if not message.text.strip().isdigit():
            bot.send_message(message.chat.id, "Iltimos, faqat raqam kiriting. Qator raqami:")
            return
        data["row"] = int(message.text.strip())
        state["step"] = "qty"
        bot.send_message(message.chat.id, "Nechta nusxa bor? (bilmasangiz \"-\" yozing):")

    elif step == "qty":
        qty_text = message.text.strip()
        data["qty"] = int(qty_text) if qty_text.isdigit() else None
        data["added_at"] = datetime.utcnow().isoformat()

        conn = get_db()
        conn.execute(
            "INSERT INTO books (title, author, shelf, row_num, qty, added_at) VALUES (?, ?, ?, ?, ?, ?)",
            (data["title"], data["author"], data["shelf"], data["row"], data["qty"], data["added_at"]),
        )
        conn.commit()
        conn.close()

        user_states.pop(uid, None)
        preview = f"📖 <b>{data['title']}</b>\n"
        if data["author"]:
            preview += f"✍️ Muallif: {data['author']}\n"
        preview += f"📍 Javon: <b>{data['shelf']}</b>  |  Qator: <b>{data['row']}</b>"
        if data["qty"]:
            preview += f"\n🔢 Soni: {data['qty']}"

        bot.send_message(
            message.chat.id,
            "✅ Kitob bazaga qo'shildi!\n\n" + preview,
            parse_mode="HTML",
            reply_markup=main_menu_keyboard(uid),
        )


# ============ ADMIN: Excel orqali ommaviy yuklash ============

@bot.message_handler(func=lambda m: m.text == "📂 Excel yuklash")
def excel_upload_prompt(message):
    if not is_admin(message.from_user.id):
        return
    text = (
        "📂 Excel (.xlsx) fayl yuboring.\n\n"
        "Ustunlar tartibi (1-qator sarlavha bo'lsin):\n"
        "<b>A:</b> Nomi | <b>B:</b> Muallif | <b>C:</b> Javon | <b>D:</b> Qator | <b>E:</b> Soni (ixtiyoriy)"
    )
    bot.send_message(message.chat.id, text, parse_mode="HTML")


@bot.message_handler(content_types=["document"])
def handle_document(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "Bu funksiya faqat adminlar uchun.")
        return

    file_name = message.document.file_name or ""
    if not file_name.lower().endswith((".xlsx", ".xlsm")):
        bot.send_message(message.chat.id, "Iltimos, .xlsx formatidagi fayl yuboring.")
        return

    file_info = bot.get_file(message.document.file_id)
    downloaded = bot.download_file(file_info.file_path)

    try:
        wb = load_workbook(io.BytesIO(downloaded), data_only=True)
        ws = wb.active
    except Exception as e:
        bot.send_message(message.chat.id, f"Faylni ochib bo'lmadi: {e}")
        return

    added = 0
    skipped = 0
    to_insert = []

    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        if not row_cells or not row_cells[0]:
            skipped += 1
            continue
        title = str(row_cells[0]).strip()
        author = str(row_cells[1]).strip() if len(row_cells) > 1 and row_cells[1] else ""
        shelf_raw = row_cells[2] if len(row_cells) > 2 else None
        row_raw = row_cells[3] if len(row_cells) > 3 else None
        qty_raw = row_cells[4] if len(row_cells) > 4 else None

        try:
            shelf = int(shelf_raw)
            row_num = int(row_raw)
        except (TypeError, ValueError):
            skipped += 1
            continue

        qty = None
        try:
            qty = int(qty_raw) if qty_raw not in (None, "") else None
        except (TypeError, ValueError):
            qty = None

        to_insert.append((title, author, shelf, row_num, qty, datetime.utcnow().isoformat()))
        added += 1

    if to_insert:
        conn = get_db()
        conn.executemany(
            "INSERT INTO books (title, author, shelf, row_num, qty, added_at) VALUES (?, ?, ?, ?, ?, ?)",
            to_insert,
        )
        conn.commit()
        conn.close()

    bot.send_message(
        message.chat.id,
        f"✅ Yuklandi: {added} ta kitob qo'shildi.\n⚠️ O'tkazib yuborildi: {skipped} ta qator (noto'g'ri format).",
        reply_markup=main_menu_keyboard(message.from_user.id),
    )


# ============ ADMIN: Javon bo'yicha ko'rish ============

@bot.message_handler(func=lambda m: m.text == "📋 Javon bo'yicha ko'rish")
def shelf_view_prompt(message):
    if not is_admin(message.from_user.id):
        return
    user_states[message.from_user.id] = {"step": "shelf_view", "data": {}}
    bot.send_message(message.chat.id, "Qaysi javon raqamini ko'rmoqchisiz?", reply_markup=cancel_keyboard())


@bot.message_handler(func=lambda m: user_states.get(m.from_user.id, {}).get("step") == "shelf_view")
def handle_shelf_view(message):
    uid = message.from_user.id
    if not message.text.strip().isdigit():
        bot.send_message(message.chat.id, "Faqat raqam kiriting:")
        return
    shelf_num = int(message.text.strip())
    user_states.pop(uid, None)

    conn = get_db()
    results = conn.execute(
        "SELECT * FROM books WHERE shelf = ? ORDER BY row_num", (shelf_num,)
    ).fetchall()
    conn.close()

    if not results:
        bot.send_message(message.chat.id, f"{shelf_num}-javonda kitob topilmadi.", reply_markup=main_menu_keyboard(uid))
        return

    grouped = {}
    for b in results:
        grouped.setdefault(b["row_num"], []).append(b)

    text = f"📚 <b>{shelf_num}-javon</b>\n\n"
    for row_num in sorted(grouped.keys()):
        text += f"<b>Qator {row_num}:</b>\n"
        for b in grouped[row_num]:
            text += f"  • {b['title']}"
            if b["author"]:
                text += f" ({b['author']})"
            text += "\n"
        text += "\n"

    bot.send_message(message.chat.id, text, parse_mode="HTML", reply_markup=main_menu_keyboard(uid))


# ============ ADMIN: Statistika ============

@bot.message_handler(func=lambda m: m.text == "📊 Statistika")
def show_stats(message):
    if not is_admin(message.from_user.id):
        return
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) AS c FROM books").fetchone()["c"]
    shelves = conn.execute("SELECT COUNT(DISTINCT shelf) AS c FROM books").fetchone()["c"]
    conn.close()
    text = f"📊 Jami kitoblar: <b>{total}</b>\n📚 Javonlar soni: <b>{shelves}</b>"
    bot.send_message(message.chat.id, text, parse_mode="HTML")


# ============ FOYDALANUVCHI: Kitob qidirish ============

@bot.message_handler(func=lambda m: m.text == "🔍 Kitob qidirish")
def search_prompt(message):
    bot.send_message(message.chat.id, "Kitob nomini yoki muallifni yozing:")


@bot.message_handler(func=lambda m: True, content_types=["text"])
def handle_search(message):
    # Admin oqimlari yuqorida ushlanadi; bu yerga faqat oddiy matn qidiruv sifatida keladi
    query = message.text.strip()
    if not query:
        return

    results = search_books(query, limit=8)

    if not results:
        bot.send_message(
            message.chat.id,
            f"❌ \"{query}\" bo'yicha hech narsa topilmadi.\nBoshqa nom yoki muallif bilan urinib ko'ring.",
        )
        return

    if len(results) == 1:
        bot.send_message(message.chat.id, format_book(results[0]), parse_mode="HTML")
    else:
        text = f"🔎 \"{query}\" bo'yicha {len(results)} ta natija topildi:\n\n"
        for b in results:
            text += format_book(b) + "\n\n"
        bot.send_message(message.chat.id, text, parse_mode="HTML")


# ============ FLASK WEBHOOK ============

@app.route("/", methods=["GET"])
def index():
    return "Kutubxona bot ishlayapti ✅"


@app.route(f"/webhook/{BOT_TOKEN}", methods=["POST"])
def webhook():
    json_data = request.get_data().decode("utf-8")
    update = telebot.types.Update.de_json(json_data)
    bot.process_new_updates([update])
    return "OK", 200


@app.route("/set_webhook", methods=["GET"])
def set_webhook():
    bot.remove_webhook()
    success = bot.set_webhook(url=f"{WEBHOOK_URL}/webhook/{BOT_TOKEN}")
    return f"Webhook o'rnatildi: {success}"


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
